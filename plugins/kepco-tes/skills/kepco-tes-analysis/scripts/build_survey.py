#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""현장 목록 정리표 빌더 — 여러 설치계획서를 한눈에 비교하는 표를 만든다.

  python build_survey.py survey_{주제}.json [-o 출력.xlsx]

용도: 사외전문가 활동처럼 **여러 현장을 일괄 검토**할 때, 자료 유무·수록 위치·검증 결과를
현장별 한 행으로 정리한다. 판독값은 JSON에, 표 서식·집계 수식은 이 스크립트가 소유한다(R8).

JSON 구조
{
 "output": "...xlsx",
 "sheets": [{
   "sheet_name": "...", "title": "...", "note": "표 읽는 법(선택)",
   "columns": [{"name":"현장","width":34}, {"name":"유무","width":9,"align":"center",
                "highlight":{"유":"ok","무":"bad","부분":"warn"}}, ...],
   "rows": [["200115 공주애터미", "유", ...], ...],
   "summary": [["총 현장","=COUNTA(B5:B30)"], ["유","=COUNTIF(C5:C30,\\"유\\")"]]
 }]
}
- `highlight`: 셀 값에 따라 음영 (ok=연녹색 / warn=연노랑 / bad=연빨강)
- `summary`: 데이터 아래 한 줄 띄우고 `항목 | 값(수식 가능)` 2열로 기재
- 저장은 _xlsx_det로 정규화해 **재실행 시 바이트 동일**
"""
import argparse
import json
import sys

from openpyxl import Workbook

from _xlsx_det import save_deterministic, sha256
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

TITLE_F = Font(bold=True, size=13)
HDR_F = Font(bold=True, color="FFFFFF")
HDR_B = PatternFill("solid", fgColor="4472C4")
FILLS = {"ok": PatternFill("solid", fgColor="E2EFDA"),
         "warn": PatternFill("solid", fgColor="FFF2CC"),
         "bad": PatternFill("solid", fgColor="FCE4E4")}
NOTE_F = Font(size=9, italic=True, color="666666")
_thin = Side(style="thin", color="B0B0B0")
BOX = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

HDR_ROW, DATA_ROW = 4, 5


def build(ws, sh):
    ws["A1"] = sh["title"]
    ws["A1"].font = TITLE_F
    if sh.get("note"):
        ws["A2"] = sh["note"]
        ws["A2"].font = NOTE_F

    cols = sh["columns"]
    for c, col in enumerate(cols, 1):
        cell = ws.cell(HDR_ROW, c, col["name"])
        cell.font, cell.fill, cell.border = HDR_F, HDR_B, BOX
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    r = DATA_ROW
    for row in sh["rows"]:
        for c, (col, v) in enumerate(zip(cols, row), 1):
            cell = ws.cell(r, c, v)
            cell.border = BOX
            cell.alignment = Alignment(vertical="top", wrap_text=True,
                                       horizontal=col.get("align", "left"))
            key = (col.get("highlight") or {}).get(str(v))
            if key:
                cell.fill = FILLS[key]
                cell.font = Font(bold=True)
        r += 1
    last = r - 1

    summary = sh.get("summary") or []
    if summary:
        r += 1                                   # 데이터와 한 줄 띄움
        ws.cell(r, 1, "집계").font = Font(bold=True)
        for label, value in summary:
            ws.cell(r, 2, label)
            ws.cell(r, 3, value)
            r += 1

    for c, col in enumerate(cols, 1):
        ws.column_dimensions[chr(ord("A") + c - 1)].width = col.get("width", 18)
    ws.freeze_panes = "A%d" % DATA_ROW
    ws.auto_filter.ref = "A%d:%s%d" % (HDR_ROW, chr(ord("A") + len(cols) - 1), last)
    return last


def main():
    ap = argparse.ArgumentParser(description="현장 목록 정리표 빌더")
    ap.add_argument("spec")
    ap.add_argument("-o", "--output")
    args = ap.parse_args()

    with open(args.spec, encoding="utf-8") as f:
        spec = json.load(f)
    out = args.output or spec.get("output") or "정리표.xlsx"

    wb = Workbook()
    for i, sh in enumerate(spec["sheets"]):
        ws = wb.active if i == 0 else wb.create_sheet()
        ws.title = sh["sheet_name"]
        n = build(ws, sh)
        print("  %-24s %d행" % (ws.title, n))

    save_deterministic(wb, out)
    print("saved:", out)
    print("sha256:", sha256(out))
    return 0


if __name__ == "__main__":
    sys.exit(main())
