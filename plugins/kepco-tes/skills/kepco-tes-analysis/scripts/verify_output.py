#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""분석 산출물(12~16시트 단일 엑셀) 규격 검사기 — SKILL.md '완성 후 필수 검증' 자동화

  python verify_output.py "KEPCO_TES_installation_plan_analysis_(파일명).xlsx"

현장마다 설비 구성이 다르므로(공랭식 → 냉각탑 없음, 냉열전용 → 난방표 미수록 등)
**시트 목록을 하드코딩하지 않는다.** 필수 시트만 확인하고, 선택 시트는 있으면 검사한다.
열 위치도 헤더 문구로 자동 탐지한다(셀맵이 개정돼도 따라간다).

하나라도 실패하면 종료코드 1 — 빌드 파이프라인의 마지막 단계로 쓴다.
"""
import argparse
import re
import sys

import openpyxl

# ── 표준 시트 순서 (_xlsx_det.SHEET_ORDER 와 동일 규격) ─────────────────
SHEET_ORDER = [
    "업체확인필요사항", "체크리스트", "상호검증", "고객현황", "건물부하계산서", "실별부하집계",
    "건물현황및냉방부하현황", "축냉설비설치계획", "감소전력계산",
    "시간대별운전계획(냉방)", "시간대별운전계획(난방)",
    "열원기기", "축열조", "냉각탑", "펌프", "열교환기", "링블로워",
]
# 설비 구성과 무관하게 항상 있어야 하는 시트
MANDATORY = [s for s in SHEET_ORDER if s not in ("실별부하집계", "냉각탑", "링블로워")]

A1_TITLE = {
    "업체확인필요사항": "설치계획서 확인사항",
    "체크리스트": "KEPCO 심야전력 축냉설비 설치계획서 검토 체크리스트",
    "상호검증": "설치계획서 내부 상호검증 (중복 기재 수치 대조)",
    "고객현황": "가. 고객현황",
    "건물부하계산서": "건물부하계산서 검토",
    "건물현황및냉방부하현황": "나. 건물현황 및 냉방부하 현황",
    "축냉설비설치계획": "다. 축냉설비 설치계획",
    "감소전력계산": "감소전력 산정 (수요관리업무처리지침 제2장 1.나)",
    "시간대별운전계획(냉방)": "시간대별 운전계획(냉방)",
    "시간대별운전계획(난방)": "시간대별 운전계획(난방)",
    "열원기기": "열원기기", "축열조": "축열조", "냉각탑": "냉각탑",
    "펌프": "펌프 순환유량(LPM/RT) 검증", "열교환기": "열교환기", "링블로워": "링블로워",
}
# 설비·현황 시트는 B1에 수록 위치를 갖는다 (`수록:` — 지침이 근거인 시트는 `근거:`)
NEEDS_B1 = ["고객현황", "건물현황및냉방부하현황", "축냉설비설치계획", "감소전력계산",
            "열원기기", "축열조", "냉각탑", "펌프", "열교환기", "링블로워"]

RESULTS = []


def chk(code, desc, ok, detail=""):
    RESULTS.append((code, desc, bool(ok), detail))


def find_header_col(ws, keyword, rows=(2, 3, 10, 54)):
    """헤더 행에서 keyword가 든 열 번호를 찾는다(셀맵 개정에 따라가기 위함)."""
    for r in rows:
        for c in range(1, ws.max_column + 1):
            if keyword in str(ws.cell(r, c).value or ""):
                return r, c
    return None, None


def is_text_stored_as_formula(v):
    """'='로 시작하지만 실제 수식이 아닌 셀 — 엑셀이 '제거된 레코드'로 지워버린다."""
    if not (isinstance(v, str) and v.startswith("=")):
        return False
    if re.match(r"^=\s", v):                    # '= 냉방에…' 처럼 등호 뒤 공백
        return True
    t = re.sub(r'"[^"]*"', "", v)               # 큰따옴표 = 문자열 리터럴
    t = re.sub(r"'[^']*'", "", t)               # 작은따옴표 = 시트명 참조
    return bool(re.search(r"[가-힣]", t))       # 남은 곳에 한글 → 수식 아님


def main():
    ap = argparse.ArgumentParser(description="KEPCO TES 분석 산출물 규격 검사")
    ap.add_argument("xlsx")
    ap.add_argument("-q", "--quiet", action="store_true", help="실패 항목만 출력")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.xlsx)
    names = wb.sheetnames
    present = [n for n in SHEET_ORDER if n in names]

    # 1 시트 구성·순서
    missing = [n for n in MANDATORY if n not in names]
    extra = [n for n in names if n not in SHEET_ORDER]
    order_ok = [n for n in names if n in SHEET_ORDER] == present
    chk("1", "시트 구성·순서", not missing and not extra and order_ok,
        "누락 %s / 미등록 %s / 순서 %s" % (missing or "-", extra or "-", "OK" if order_ok else "어긋남"))

    # 2 업체확인필요사항 = 맨 앞·노란 탭·4열 양식·데이터 없음
    q = wb["업체확인필요사항"]
    tab = str(getattr(q.sheet_properties.tabColor, "rgb", "") or "")
    hdr4 = [q.cell(3, c).value for c in range(1, 5)]
    empty = all(q.cell(r, c).value is None for r in range(4, 60) for c in range(1, 8))
    chk("2", "업체확인필요사항 양식", names[0] == "업체확인필요사항" and tab.endswith("FFFF00")
        and hdr4 == ["번호", "구분(시트/단계)", "확인필요항목 및 내용", "업체 답변"] and empty,
        "헤더 %s / 행4~ 비어있음 %s" % (hdr4, empty))

    # 3 A1 제목
    bad = [n for n in present if A1_TITLE.get(n) and wb[n]["A1"].value != A1_TITLE[n]]
    chk("3", "A1 제목", not bad, str(bad))

    # 4 체크리스트 판정·근거 위치
    c = wb["체크리스트"]
    e = [c.cell(r, 5).value for r in range(7, 28)]
    blank_e = [r for r in range(7, 28) if c.cell(r, 5).value in (None, "")]
    blank_f = [r for r in range(7, 28) if not c.cell(r, 6).value]
    chk("4", "체크리스트 행7~27 판정·근거", not blank_e and not blank_f,
        "✅%d ⚠️%d ❌%d 해당없음%d | 판정 빈칸 %s | 근거 빈칸 %s"
        % (e.count("✅"), e.count("⚠️"), e.count("❌"), e.count("해당없음"),
           blank_e or "-", blank_f or "-"))

    # 5 건물부하계산서 — '무'면 행8 이하 공란
    lc = wb["건물부하계산서"]
    유무 = lc["B4"].value
    tail_empty = all(lc.cell(r, cc).value is None for r in range(8, 30) for cc in range(1, 6))
    chk("5", "건물부하계산서 유무 처리", 유무 == "유" or tail_empty,
        "유무=%s / 행8이하 공란=%s" % (유무, tail_empty))

    # 6 시간대별운전계획 — 누계 점화식·V1~V14
    op = wb["시간대별운전계획(냉방)"]
    f11, f12 = str(op["F11"].value or ""), str(op["F12"].value or "")
    vcodes = all(op.cell(r, 1).value == "V%d" % (r - 38) for r in range(39, 53))
    chk("6", "시간대별운전계획 누계·V1~V14",
        f11 == "=C11-E11" and f12 == "=F11+C12-E12" and vcodes,
        "F11=%s F12=%s V1~V14=%s" % (f11, f12, vcodes))

    # 7 열원기기 — 단위환산 검증 + 모드별 용량 적정성
    h = wb["열원기기"]
    conv = [str(h.cell(r, 4).value or "") for r in range(17, 21)]
    cap = all(str(h.cell(r, 2).value or "").startswith("=IF(") for r in (46, 48))
    chk("7", "열원기기 환산검증·용량 적정성",
        all(x.startswith("=") for x in conv) and cap, "D17~D20=%s / 판정수식=%s" % (conv, cap))

    # 8 축열조 — 비교검증 앵커 + 계산자료 ①~⑥
    t = wb["축열조"]
    anchors = all(t["B%d" % r].value is not None for r in (56, 57, 59, 60))
    marks = [str(t.cell(r, 1).value or "")[:1] for r in range(63, 69)]
    chk("8", "축열조 비교검증·계산자료", anchors and marks == list("①②③④⑤⑥"),
        "B56/57/59/60=%s / 계산자료=%s" % (anchors, "".join(marks)))

    # 9 펌프 — 블록마다 LPM/RT 수식·판정 수식
    p = wb["펌프"]
    starts = [r for r in range(1, p.max_row + 1) if str(p.cell(r, 1).value or "").startswith("[펌프 ")]
    pump_ok = starts and all(str(p.cell(s + 7, 2).value or "").startswith("=B")
                             and str(p.cell(s + 9, 2).value or "").startswith("=IF(AND(")
                             for s in starts)
    chk("9", "펌프 LPM/RT 수식", pump_ok, "펌프 %d대 블록" % len(starts))

    # 10 상호검증 — 판정 수식·허용오차
    x = wb["상호검증"]
    rows = [r for r in range(4, x.max_row + 1)
            if x.cell(r, 2).value and isinstance(x.cell(r, 3).value, (int, float))]
    judge_ok = all(str(x.cell(r, 10).value or "").startswith("=IF(COUNT") for r in rows)
    chk("10", "상호검증 판정 수식", rows and judge_ok, "%d개 항목" % len(rows))

    # 11 감소전력계산 — MIN 채택·판정
    g = wb["감소전력계산"]
    has_min = any("MIN(" in str(g.cell(r, 2).value or "") for r in range(1, g.max_row + 1))
    has_judge = any(str(g.cell(r, 2).value or "").startswith("=IF(ABS(")
                    for r in range(1, g.max_row + 1))
    chk("11", "감소전력 MIN 채택·기재값 대조", has_min and has_judge,
        "MIN=%s 판정=%s" % (has_min, has_judge))

    # 12 설비 목록 vs 설비 시트 대조 블록
    inst = wb["축냉설비설치계획"]
    hdr_r, _ = find_header_col(inst, "대응 시트", rows=range(50, 60))
    chk("12", "설비 목록 누락 검증 블록", hdr_r is not None, "헤더 행 %s" % hdr_r)

    # 13 숫자 서식 — '.#' 금지 (엑셀이 '900.' 처럼 소수점만 남긴다)
    dot = [(w.title, cc.coordinate) for w in wb.worksheets for row in w.iter_rows() for cc in row
           if isinstance(cc.number_format, str) and cc.number_format.endswith(".#")]
    chk("13", "'.#' 서식 없음", not dot, str(dot[:3]))

    # 14 kcal 단위 셀은 소수점 없이
    kbad = [(w.title, cc.coordinate) for w in wb.worksheets for row in w.iter_rows() for cc in row
            if cc.value is not None
            and "kcal" in "%s %s" % (w.cell(cc.row, 1).value or "", w.cell(cc.row, 3).value or "")
            and isinstance(cc.number_format, str) and "." in cc.number_format]
    chk("14", "kcal 셀 소수점 없음", not kbad, str(kbad[:3]))

    # 15 수식 유효성 — 표시용 문자열이 '='로 시작하면 엑셀이 레코드를 제거한다
    bf = [(w.title, cc.coordinate, str(cc.value)[:40]) for w in wb.worksheets
          for row in w.iter_rows() for cc in row if is_text_stored_as_formula(cc.value)]
    chk("15", "잘못된 수식 없음", not bf, str(bf[:3]))

    # 16 수록 위치 — B1 + 출처 열(헤더 자동 탐지) + 출처 열 '=' 시작 금지
    nob1 = [n for n in NEEDS_B1 if n in names
            and not str(wb[n]["B1"].value or "").startswith(("수록:", "근거:"))]
    nosrc, srcformula = [], []
    for n in present:
        if n in ("업체확인필요사항", "체크리스트", "상호검증",
                 "건물부하계산서", "시간대별운전계획(냉방)", "시간대별운전계획(난방)"):
            continue
        hr, hc = find_header_col(wb[n], "출처")
        if hc is None:
            nosrc.append(n)
            continue
        for r in range(hr + 1, wb[n].max_row + 1):
            if str(wb[n].cell(r, hc).value or "").startswith("="):
                srcformula.append((n, wb[n].cell(r, hc).coordinate))
    chk("16", "수록 위치(B1·출처 열)", not nob1 and not nosrc and not srcformula,
        "B1누락 %s / 출처열없음 %s / 수식저장 %s" % (nob1 or "-", nosrc or "-", srcformula[:3] or "-"))

    # ── 출력 ──
    fails = [r for r in RESULTS if not r[2]]
    for code, desc, ok, detail in RESULTS:
        if args.quiet and ok:
            continue
        print("%-3s %-28s %s  %s" % (code, desc, "✅" if ok else "❌", detail))
    print("\n  시트 %d개: %s" % (len(names), " / ".join(names)))
    print("  검사 %d개 중 통과 %d / 실패 %d" % (len(RESULTS), len(RESULTS) - len(fails), len(fails)))
    if fails:
        print("\n❌ 실패: " + ", ".join("%s %s" % (c, d) for c, d, _, _ in fails))
        return 1
    print("  → 전 항목 통과.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
