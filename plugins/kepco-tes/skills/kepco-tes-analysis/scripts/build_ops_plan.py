#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""시간대별 운전계획 검증 엑셀 빌더 (Phase 4 / v14 셀맵)

  python build_ops_plan.py ops_{현장}.json [-o 출력.xlsx]

원칙
- **단위 무변환**: 설치계획서에 적힌 단위(RT/USRt/RTh/kW/kWh/kcal/h)를 그대로 쓴다. 환산하지 않는다.
- **축열조 누계의 기준시각은 '축열 시작 직전'** — 그 시점을 0으로 두고 시간 순으로 누적한다.
  표가 1:00부터 시작해도 축열 시작 시각 기준으로 행을 재배열해 누계를 다시 쌓는다.
- 누계 점화식: 누계(t) = 누계(t-1) + 열원기기 축열량(t) − 축열조 방열량(t)

JSON 스키마: assets/opsplan_schema.example.json
빌드 후 파이썬에서도 동일 검증을 돌려 결과를 보고한다(불일치 시 종료코드 1).
"""
import argparse
import json
import re
import sys

from openpyxl import Workbook

from _xlsx_det import save_deterministic, sha256, open_or_new, replace_sheet, order_sheets
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

TITLE_F = Font(bold=True, size=13)
H1_F = Font(bold=True, size=11)
HDR_F = Font(bold=True, color="FFFFFF")
HDR_B = PatternFill("solid", fgColor="4472C4")
SUM_B = PatternFill("solid", fgColor="D9E1F2")
CHG_B = PatternFill("solid", fgColor="E2EFDA")   # 축열(충전) 시간대
DIS_B = PatternFill("solid", fgColor="FFF2CC")   # 방열 시간대
NOTE_F = Font(size=9, color="666666")
_thin = Side(style="thin", color="B0B0B0")
BOX = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
NA = "해당없음"
MISS = "— (미기재)"


def build_absent(ws, spec):
    """운전계획표가 문서에 없는 경우 — [1] 블록만 쓰고 끝낸다 (phase4 §(1)).

    표를 지어내지 않으면서도 12시트 구성은 유지한다. 냉방표만 있고 난방표가 없는
    문서에서 `시간대별운전계획(난방)` 시트를 이 형태로 만든다.
    """
    ws["A1"] = spec.get("title", "시간대별 운전계획 검증 (난방)")
    ws["A1"].font = TITLE_F
    ws["A3"] = "[1] 수록 위치 및 표 개요"
    ws["A3"].font = H1_F
    for i, (k, v) in enumerate([("수록 위치", MISS), ("표 제목", NA),
                                ("단위", NA), ("축열 시작 시각", NA)]):
        ws.cell(4 + i, 1, k).font = Font(bold=True)
        ws.cell(4 + i, 2, v)
    for col, w in zip("ABCDEFGHI", (12, 40, 18, 16, 46, 18, 14, 16, 14)):
        ws.column_dimensions[col].width = w
    return []


def _header(ws, row, labels):
    for c, v in enumerate(labels, 1):
        cell = ws.cell(row, c, v)
        cell.font, cell.fill, cell.border = HDR_F, HDR_B, BOX
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


NIGHT_HOURS = {22, 23, 0, 1, 2, 3, 4, 5, 6, 7}   # 심야전력 시간대 22:00~08:00 (시작시 기준)


def night_labels(rows, spec):
    """심야 시간대 행의 시각 라벨을 정한다.

    spec['심야_시각']이 있으면 그대로 쓰고, 없으면 라벨에서 자동 판별한다:
      - 구간 표기(`24:00-01:00`, `0-1`, `22:00~23:00`) → 앞 숫자를 **시작시**로 본다
      - 단일 표기(`23:00`, `8:00`) → 그 시각을 **종료시**로 보아 시작시 = t-1
    자동 판별 결과는 표준출력에 찍어 사람이 검토할 수 있게 한다.
    """
    if spec.get("심야_시각"):
        return list(spec["심야_시각"]), "JSON 지정"
    out = []
    for row in rows:
        lab = str(row["시각"])
        nums = re.findall(r"\d+", lab)
        if not nums:
            continue
        start = int(nums[0]) % 24 if re.search(r"[-~]", lab) else (int(nums[0]) - 1) % 24
        if start in NIGHT_HOURS:
            out.append(row["시각"])
    return out, "라벨 자동판별"


def reorder(rows, start_hour):
    """축열 시작 직전이 첫 행이 되도록 24시간 표를 회전한다."""
    idx = next((i for i, r in enumerate(rows) if r["시각"] == start_hour), None)
    if idx is None:
        return rows
    return rows[idx:] + rows[:idx]


def build(ws, spec):
    src = spec["source"]
    u_load = spec["units"]["부하"]            # 예: USRt, RT, kW
    u_cum = spec["units"]["누계"]             # 예: USRt·h, RTh, kWh
    rows = reorder(spec["rows"], spec["축열_시작시각"])

    ws["A1"] = spec.get("title", "시간대별 운전계획 검증 (냉방)")
    ws["A1"].font = TITLE_F

    # [1] 수록 위치
    ws["A3"] = "[1] 수록 위치 및 표 개요"
    ws["A3"].font = H1_F
    meta = [
        ("수록 위치", src["위치"]),
        ("표 제목", src.get("표제목", NA)),
        ("단위", "부하·운전량 %s / 누계 %s" % (u_load, u_cum)),
        ("축열 시작 시각", "%s (누계 기준시각 = 축열 시작 직전)" % spec["축열_시작시각"]),
    ]
    for i, (k, v) in enumerate(meta):
        ws.cell(4 + i, 1, k).font = Font(bold=True)
        ws.cell(4 + i, 2, v)

    # [2] 시간대별 표
    ws["A9"] = "[2] 시간대별 부하·운전계획·축열조 누계"
    ws["A9"].font = H1_F
    _header(ws, 10, [
        "시각", "총부하(%s)" % u_load, "열원기기 축열운전(%s)" % u_load,
        "열원기기 직접담당(%s)" % u_load, "축열조 방열(%s)" % u_load,
        "축열조 누계(%s, 수식)" % u_cum, "부하 정합(수식)",
        "계획서 기재 누계", "누계 대조(수식)",
    ])
    R0 = 11
    r = R0
    for i, row in enumerate(rows):
        ws.cell(r, 1, row["시각"])
        ws.cell(r, 2, row.get("총부하", 0))
        ws.cell(r, 3, row.get("축열운전", 0))
        ws.cell(r, 4, row.get("직접담당", 0))
        ws.cell(r, 5, row.get("방열", 0))
        # 누계: 기준시각(첫 행) 직전을 0으로 두고 점화
        ws.cell(r, 6, "=C{0}-E{0}".format(r) if i == 0 else "=F{0}+C{1}-E{1}".format(r - 1, r))
        ws.cell(r, 7, '=IF(ABS(B{0}-(D{0}+E{0}))<=0.5,"✅","⚠️")'.format(r))
        ws.cell(r, 8, row.get("기재누계", NA))
        ws.cell(r, 9, ('=IF(ABS(F{0}-H{0})<=0.5,"✅","⚠️")'.format(r)
                       if row.get("기재누계") is not None else NA))
        for c in range(1, 10):
            ws.cell(r, c).border = BOX
        for c in (2, 3, 4, 5, 6, 8):
            ws.cell(r, c).number_format = "#,##0.0"
        if row.get("축열운전"):
            for c in range(1, 10):
                ws.cell(r, c).fill = CHG_B
        elif row.get("방열"):
            for c in range(1, 10):
                ws.cell(r, c).fill = DIS_B
        r += 1
    R1 = r - 1

    ws.cell(r, 1, "합계").font = Font(bold=True)
    for c in (2, 3, 4, 5):
        col = chr(ord("A") + c - 1)
        ws.cell(r, c, "=SUM({0}{1}:{0}{2})".format(col, R0, R1))
        ws.cell(r, c).number_format = "#,##0.0"
    for c in range(1, 10):
        ws.cell(r, c).border = BOX
        ws.cell(r, c).fill = SUM_B
    R_SUM = r

    # [3] 검증
    r += 2
    ws.cell(r, 1, "[3] 검증").font = H1_F
    r += 1
    _header(ws, r, ["코드", "검증 내용", "계산값(수식)", "대조값", "판정"])
    r += 1
    tot = spec.get("기재합계", {})
    cap = spec.get("축열조_설계용량")
    checks = [
        ("V1", "총부하 합계 = 계획서 기재 합계", "=B%d" % R_SUM, tot.get("총부하"),
         '=IF(ABS(C{r}-D{r})<=1,"✅","⚠️")'),
        ("V2", "열원기기 축열운전 합계 = 기재", "=C%d" % R_SUM, tot.get("축열운전"),
         '=IF(ABS(C{r}-D{r})<=1,"✅","⚠️")'),
        ("V3", "열원기기 직접담당 합계 = 기재", "=D%d" % R_SUM, tot.get("직접담당"),
         '=IF(ABS(C{r}-D{r})<=1,"✅","⚠️")'),
        ("V4", "축열조 방열 합계 = 기재", "=E%d" % R_SUM, tot.get("방열"),
         '=IF(ABS(C{r}-D{r})<=1,"✅","⚠️")'),
        ("V5", "축열량 합계 = 방열량 합계 (수지 균형)", "=C{0}-E{0}".format(R_SUM), 0,
         '=IF(ABS(C{r}-D{r})<=1,"✅","⚠️")'),
        ("V6", "최종 누계 = 0 (기준시각 복귀)", "=F%d" % R1, 0,
         '=IF(ABS(C{r}-D{r})<=1,"✅","⚠️")'),
        ("V7", "최대 누계 ≤ 축열조 설계용량", "=MAX(F{0}:F{1})".format(R0, R1), cap,
         '=IF(C{r}<=D{r}*1.001,"✅","❌")'),
        ("V8", "축열률 = 방열 합계 ÷ 총부하 합계 (기준 ≥40%)",
         "=E{0}/B{0}".format(R_SUM), 0.4, '=IF(C{r}>=D{r},"✅","❌")'),
        ("V9", "피크부하 = 시간대별 최대 총부하", "=MAX(B{0}:B{1})".format(R0, R1),
         spec.get("기재_피크부하"), '=IF(ABS(C{r}-D{r})<=1,"✅","⚠️")'),
        ("V10", "부하 정합(부하 = 직접담당 + 방열) 전 시각 충족",
         '=COUNTIF(G{0}:G{1},"⚠️")'.format(R0, R1), 0, '=IF(C{r}=D{r},"✅","⚠️")'),
        ("V11", "계획서 기재 누계와 재계산 누계 일치",
         '=COUNTIF(I{0}:I{1},"⚠️")'.format(R0, R1), 0, '=IF(C{r}=D{r},"✅","⚠️")'),
    ]

    # 심야시간대 검증 (KEPCO 심사항목) — spec에 심야 시각 목록이 있을 때만
    night, night_src = night_labels(rows, spec)
    night_rows = [i for i, row in enumerate(rows) if row["시각"] in night]
    spec["_심야판별"] = (night, night_src)
    if night_rows:
        # 심야 행은 회전 후 위치가 흩어질 수 있으므로 SUMIF/MAX 대신 개별 셀 참조를 합친다
        n_load = "+".join("B%d" % (R0 + i) for i in night_rows)
        n_max = ",".join("B%d" % (R0 + i) for i in night_rows)
        day_rows = [i for i in range(len(rows)) if i not in night_rows]
        d_max = ",".join("B%d" % (R0 + i) for i in day_rows)
        checks += [
            ("V12", "심야(22:00~08:00) 부하 합계 — 0이면 심야부하 없음",
             "=" + n_load, 0, '=IF(C{r}=0,"✅ 없음","있음")'),
            ("V13", "심야 최대부하 ≤ 주간 최대부하 × 60%",
             "=MAX(%s)" % n_max, "=MAX(%s)*0.6" % d_max, '=IF(C{r}<=D{r},"✅","❌")'),
            ("V14", "심야 부하량 ≤ 일일 부하량 × 40%",
             "=" + n_load, "=B%d*0.4" % R_SUM, '=IF(C{r}<=D{r},"✅","❌")'),
        ]
    else:
        for code, desc in (("V12", "심야(22:00~08:00) 부하 합계 — 0이면 심야부하 없음"),
                           ("V13", "심야 최대부하 ≤ 주간 최대부하 × 60%"),
                           ("V14", "심야 부하량 ≤ 일일 부하량 × 40%")):
            checks.append((code, desc + " (심야 행 판별 불가)", None, None, None))
    for code, desc, val, comp, judge in checks:
        if comp is None:
            val, comp, judge = NA, NA, NA
        ws.cell(r, 1, code)
        ws.cell(r, 2, desc)
        ws.cell(r, 3, val)
        ws.cell(r, 4, comp)
        ws.cell(r, 5, judge.format(r=r) if judge != NA else NA)
        for c in range(1, 6):
            ws.cell(r, c).border = BOX
        if code == "V8":
            ws.cell(r, 3).number_format = "0.0%"
            ws.cell(r, 4).number_format = "0.0%"
        else:
            ws.cell(r, 3).number_format = "#,##0.0"
        r += 1

    # [4] 확인사항
    notes = spec.get("notes") or []
    if notes:
        r += 1
        ws.cell(r, 1, "[4] 확인사항").font = H1_F
        r += 1
        _header(ws, r, ["번호", "항목", "계획서", "검증값", "확인요청 내용"])
        r += 1
        for i, n in enumerate(notes, 1):
            ws.cell(r, 1, i)
            ws.cell(r, 2, n["항목"])
            ws.cell(r, 3, n.get("계획서", NA))
            ws.cell(r, 4, n.get("검증값", NA))
            ws.cell(r, 5, n["확인요청"])
            for c in range(1, 6):
                ws.cell(r, c).border = BOX
                ws.cell(r, c).alignment = Alignment(vertical="top", wrap_text=True)
            r += 1

    for col, w in zip("ABCDEFGHI", (12, 40, 18, 16, 46, 18, 14, 16, 14)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A11"
    return rows


def verify(spec, rows):
    """엑셀 수식과 같은 검증을 파이썬으로 재수행 (빌드 단계 자체 점검)."""
    out = []
    s_load = sum(r.get("총부하", 0) for r in rows)
    s_chg = sum(r.get("축열운전", 0) for r in rows)
    s_dir = sum(r.get("직접담당", 0) for r in rows)
    s_dis = sum(r.get("방열", 0) for r in rows)
    tot = spec.get("기재합계", {})

    def add(code, val, comp, ok=None):
        if comp is None:
            return
        out.append((code, val, comp, (abs(val - comp) <= 1) if ok is None else ok))

    add("V1 총부하 합계", s_load, tot.get("총부하"))
    add("V2 축열운전 합계", s_chg, tot.get("축열운전"))
    add("V3 직접담당 합계", s_dir, tot.get("직접담당"))
    add("V4 방열 합계", s_dis, tot.get("방열"))
    add("V5 축열−방열", s_chg - s_dis, 0)

    cum, mx, mismatch_load, mismatch_cum = 0.0, 0.0, 0, 0
    for r in rows:
        cum += r.get("축열운전", 0) - r.get("방열", 0)
        mx = max(mx, cum)
        if abs(r.get("총부하", 0) - (r.get("직접담당", 0) + r.get("방열", 0))) > 0.5:
            mismatch_load += 1
        if r.get("기재누계") is not None and abs(cum - r["기재누계"]) > 0.5:
            mismatch_cum += 1
    add("V6 최종 누계", cum, 0)
    if spec.get("축열조_설계용량"):
        out.append(("V7 최대 누계 ≤ 설계용량", mx, spec["축열조_설계용량"],
                    mx <= spec["축열조_설계용량"] * 1.001))
    if s_load:
        ratio = s_dis / s_load * 100
        out.append(("V8 축열률(%)", round(ratio, 1), 40.0, ratio >= 40))
    add("V9 피크부하", max((r.get("총부하", 0) for r in rows), default=0),
        spec.get("기재_피크부하"))
    out.append(("V10 부하 정합 불일치 행수", mismatch_load, 0, mismatch_load == 0))
    out.append(("V11 기재누계 불일치 행수", mismatch_cum, 0, mismatch_cum == 0))

    night = set(spec.get("_심야판별", ([], ""))[0])
    if night:
        nl = [r.get("총부하", 0) or 0 for r in rows if r["시각"] in night]
        dl = [r.get("총부하", 0) or 0 for r in rows if r["시각"] not in night]
        n_sum, n_max, d_max = sum(nl), max(nl, default=0), max(dl, default=0)
        out.append(("V13 심야최대 ≤ 주간최대×60%", n_max, round(d_max * 0.6, 1), n_max <= d_max * 0.6))
        out.append(("V14 심야부하량 ≤ 일일×40%", n_sum, round(s_load * 0.4, 1), n_sum <= s_load * 0.4))
    return out


def main():
    ap = argparse.ArgumentParser(description="시간대별 운전계획 검증 엑셀 빌더 (Phase 4)")
    ap.add_argument("spec")
    ap.add_argument("-o", "--output")
    ap.add_argument("--into", help="기존 분석 워크북에 시트를 끼워 넣는다 (단일 파일 산출 — v17 기본)")
    args = ap.parse_args()

    with open(args.spec, encoding="utf-8") as f:
        spec = json.load(f)

    if args.into:                        # 단일 파일 모드 — 12시트 워크북에 병합
        out = args.into
        wb = open_or_new(out)
    else:                                # 단독 검증 파일 모드
        out = args.output or spec.get("output") or "시간대별운전계획.xlsx"
        wb = Workbook()
        wb.remove(wb.active)

    sheets = spec["sheets"] if "sheets" in spec else [spec]
    all_checks = []
    for sh in sheets:
        ws = replace_sheet(wb, sh.get("sheet_name", "시간대별운전계획(냉방)"))
        if sh.get("미수록"):             # 문서에 표가 없음 — [1]만 쓰고 끝낸다
            build_absent(ws, sh)
            print("  [%s] 표 미수록 — [1] 블록만 생성" % ws.title)
            continue
        rows = build(ws, sh)
        all_checks.append((ws.title, verify(sh, rows)))
        nl, src = sh.get("_심야판별", ([], "-"))
        print("  심야 행(%s, %d개): %s" % (src, len(nl), ", ".join(map(str, nl)) or "없음"))

    if args.into:
        order_sheets(wb)

    save_deterministic(wb, out)          # 재실행 시 바이트 동일 보장(R8)
    print("saved:", out)
    print("sha256:", sha256(out))

    bad = 0
    for title, checks in all_checks:
        print("\n[%s] 검증" % title)
        for code, val, comp, ok in checks:
            print("  %-26s 계산 %10.1f | 기준 %10.1f | %s" % (code, val, comp, "✅" if ok else "⚠️"))
            bad += (not ok)
    if bad:
        print("\n⚠️ %d개 항목 불일치 — 원문 재판독 또는 확인사항 기재가 필요하다." % bad)
        return 1
    print("\n  → 전 항목 일치.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
