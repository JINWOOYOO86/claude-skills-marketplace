# -*- coding: utf-8 -*-
"""결정론 저장 헬퍼 — 같은 입력이면 **바이트 동일**한 xlsx가 나오도록 정규화한다.

openpyxl은 저장할 때 docProps/core.xml에 현재 시각(created/modified)을 넣고
zip 멤버에도 저장 시각을 기록하기 때문에, 내용이 같아도 실행할 때마다 sha256이 달라진다.
이 모듈은 (a) 문서 속성 타임스탬프, (b) zip 멤버 시각·압축 방식을 고정값으로 정규화한다.
(R8 결정론 빌드 규칙 — assets/cases/build_case_*.py 와 같은 취지)
"""
import datetime
import re
import os
import shutil
import zipfile

EPOCH = datetime.datetime(2020, 1, 1, 0, 0, 0)   # 고정 타임스탬프
ZIP_DATE = (1980, 1, 1, 0, 0, 0)                 # zip 최소 표현 시각

# ── 단일 산출물(12시트) 표준 시트 순서 — 빌더가 기존 워크북에 끼워 넣을 때 쓴다 ──
SHEET_ORDER = [
    "업체확인필요사항",                                    # v17 — 맨 앞 (사용자 작성 양식)
    "체크리스트", "상호검증", "고객현황", "건물부하계산서", "실별부하집계", "건물현황및냉방부하현황",
    "축냉설비설치계획", "감소전력계산", "시간대별운전계획(냉방)", "시간대별운전계획(난방)",
    "열원기기", "축열조", "냉각탑", "펌프", "열교환기", "링블로워",
]


def open_or_new(path):
    """--into 대상 워크북을 연다. 없으면 빈 워크북을 새로 만든다."""
    from openpyxl import load_workbook, Workbook
    if path and os.path.exists(path):
        wb = load_workbook(path)
        if getattr(wb, "_archive", None):      # 원본 zip 핸들을 닫아야 덮어쓸 수 있다
            wb._archive.close()
        return wb
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def replace_sheet(wb, title):
    """같은 이름 시트가 있으면 지우고 새로 만든다(재실행 시 중복 방지)."""
    if title in wb.sheetnames:
        del wb[title]
    return wb.create_sheet(title)


def order_sheets(wb, aliases=None):
    """SHEET_ORDER 기준으로 시트를 정렬한다.

    aliases: {실제 시트명: 표준 시트명} — 실별부하집계처럼 이름이 가변인 시트를
    표준 위치에 놓기 위해 쓴다. 표준 목록에 없는 시트는 뒤에 원래 순서로 남는다.
    """
    aliases = aliases or {}
    names = list(wb.sheetnames)

    def key(ws):
        std = aliases.get(ws.title, ws.title)
        if std in SHEET_ORDER:
            return (0, SHEET_ORDER.index(std))
        return (1, names.index(ws.title))

    wb._sheets.sort(key=key)
    return wb.sheetnames


def save_deterministic(wb, path, creator="kepco-tes-analysis"):
    """워크북을 저장한 뒤 타임스탬프를 정규화해 재실행 시 바이트 동일을 보장한다."""
    props = wb.properties
    props.creator = creator
    props.lastModifiedBy = creator
    props.created = EPOCH
    props.modified = EPOCH
    wb.save(path)
    _normalize_zip(path)
    return path


_STAMP = EPOCH.strftime("%Y-%m-%dT%H:%M:%SZ")
_CORE_RE = re.compile(
    br'(<dcterms:(?:created|modified)[^>]*>)[^<]*(</dcterms:(?:created|modified)>)')


def _fix_core(data):
    """openpyxl이 저장 시각으로 덮어쓴 dcterms:created/modified를 고정값으로 바꾼다."""
    return _CORE_RE.sub(lambda m: m.group(1) + _STAMP.encode() + m.group(2), data)


def _normalize_zip(path):
    """zip 멤버 시각·압축방식을 고정값으로 다시 쓴다.

    임시파일 rename(os.replace)은 Windows 드라이브 마운트(DrvFS)에서 대상 파일이
    이미 존재하면 WinError 5로 거부되는 일이 있다(--into 재저장 경로). 그래서
    멤버를 전부 메모리로 읽어 원본을 닫은 뒤, 같은 경로에 그대로 덮어쓴다.
    """
    with zipfile.ZipFile(path, "r") as zin:
        members = []
        for info in zin.infolist():                 # 멤버 순서는 원본 그대로 유지
            data = zin.read(info.filename)
            if info.filename == "docProps/core.xml":
                data = _fix_core(data)
            members.append((info.filename, info.external_attr, data))

    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, attr, data in members:
            ni = zipfile.ZipInfo(name, date_time=ZIP_DATE)
            ni.compress_type = zipfile.ZIP_DEFLATED
            ni.external_attr = attr
            ni.create_system = 0
            zout.writestr(ni, data)


def sha256(path):
    import hashlib
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for b in iter(lambda: f.read(1 << 16), b""):
            h.update(b)
    return h.hexdigest()
