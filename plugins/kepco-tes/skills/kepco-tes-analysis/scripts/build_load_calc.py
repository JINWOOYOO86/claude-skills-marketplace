#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""건물부하계산서 검토 엑셀 빌더 (Phase 2 / v13 셀맵)

문서에서 판독한 값은 JSON에 두고, 셀맵·수식·검증 로직은 이 스크립트가 고정한다.
  python build_load_calc.py loadcalc_{현장}.json [-o 출력.xlsx]

JSON 스키마는 references/phase2_load_calc.md 및 assets/loadcalc_schema.example.json 참조.
빌드 후 실별 상세표의 열별 합계를 계산서 기재 Grand Total과 대조해 결과를 표준출력에 보고한다
(전 열 ✅가 아니면 종료코드 1 — 판독 오류를 빌드 단계에서 잡는다).
"""
import argparse
import json
import sys

from openpyxl import Workbook

from _xlsx_det import save_deterministic, sha256, open_or_new, replace_sheet, order_sheets
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── 고정 셀맵 (phase2_load_calc.md v13) ──────────────────────────────
R_PRESENCE = 4          # [1] 행4~7
R_AREA_HDR, R_AREA = 10, 11        # [2] 행10 헤더, 행11~16
R_LOAD_HDR, R_LOAD = 19, 20        # [3] 행19 헤더, 행20~25
R_SPACE_TITLE, R_SPACE_HDR, R_SPACE = 27, 28, 29   # [4]
SPACE_LAST = 300        # SUM 범위 하한 (데이터 행 수와 무관하게 고정)

RT_DIVISOR = {"W": 3517, "kcal/h": 3024}   # 1 RT = 3,517 W = 3,024 kcal/h
NA = "해당없음"

TITLE_F = Font(bold=True, size=13)
H1_F = Font(bold=True, size=11)
HDR_F = Font(bold=True, color="FFFFFF")
HDR_B = PatternFill("solid", fgColor="4472C4")
SUM_B = PatternFill("solid", fgColor="D9E1F2")
DUP_B = PatternFill("solid", fgColor="FFF2CC")
NOTE_F = Font(size=9, color="666666")
_thin = Side(style="thin", color="B0B0B0")
BOX = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _header(ws, row, labels):
    for c, v in enumerate(labels, 1):
        cell = ws.cell(row, c, v)
        cell.font, cell.fill, cell.border = HDR_F, HDR_B, BOX
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _row(ws, row, values, borders=True, ncols=None):
    for c, v in enumerate(values, 1):
        ws.cell(row, c, v)
    if borders:
        for c in range(1, (ncols or len(values)) + 1):
            ws.cell(row, c).border = BOX


def build_main(ws, spec, notes=True):
    """notes=False면 [5] 확인사항 블록을 생략한다.

    12시트 단일 산출물(--into)에서는 확인사항을 시트12 `업체확인필요사항`이 전담하므로
    이 시트에 중복 기재하지 않는다(phase2_load_calc.md [5]). 건물부하계산서만 단독으로
    검증하는 파일에서는 notes=True로 이 블록을 쓴다.
    """
    presence = spec["presence"]
    has_calc = presence.get("유무") == "유"
    unit = presence.get("부하 단위") or NA
    div = RT_DIVISOR.get(unit)

    ws["A1"] = spec.get("title", "건물부하계산서 검토")
    ws["A1"].font = TITLE_F

    # [1] 부하계산서 유무
    ws["A3"] = "[1] 부하계산서 유무"
    ws["A3"].font = H1_F
    for i, key in enumerate(["유무", "수록 위치", "수록 형태", "부하 단위"]):
        r = R_PRESENCE + i
        ws.cell(r, 1, "부하계산서 유무" if key == "유무" else key).font = Font(bold=True)
        ws.cell(r, 2, presence.get(key, NA))

    if not has_calc:
        # 부하계산서가 없으면 [2]~[4]는 작성하지 않는다 (억지 채움 금지)
        if notes:
            return write_notes(ws, spec, start_row=10)
        for col, w in zip("ABCDE", (30, 26, 12, 22, 46)):
            ws.column_dimensions[col].width = w
        return

    area, load = spec["area_check"], spec["load_check"]

    # [2] 면적 계산검증
    ws["A9"] = "[2] 면적 계산검증"
    ws["A9"].font = H1_F
    _header(ws, R_AREA_HDR, ["항목", "계산값(수식)", "단위", "대조값", "판정"])
    r = R_AREA
    area_rows = [
        ("공간별 냉방면적 합계", "=SUM(C{0}:C{1})".format(R_SPACE, SPACE_LAST), "㎡",
         area.get("계산서_냉방면적합계"), '=IF(ABS(B{r}-D{r})<=1,"✅","⚠️")',
         "계산서 기재 합계와 대조 (전사 누락 검증)"),
        ("공간별 난방면적 합계", "=SUM(F{0}:F{1})".format(R_SPACE, SPACE_LAST), "㎡",
         area.get("계산서_난방면적합계"), '=IF(ABS(B{r}-D{r})<=1,"✅","⚠️")',
         "계산서 기재 합계와 대조 (전사 누락 검증)"),
        ("냉방면적 − 난방면적", "=B{0}-B{1}".format(R_AREA, R_AREA + 1), "㎡", 0,
         '=IF(ABS(B{r}-D{r})<=1,"✅","⚠️")', "냉·난방 대상 면적 차이"),
        ("냉난방면적 (설치계획서 대조)", "=B{0}".format(R_AREA), "㎡",
         area.get("설치계획서_냉난방면적_m2"), '=IF(ABS(B{r}-D{r})<=1,"✅","❌")',
         "설치계획서 '나.건물개요' 냉난방면적"),
        ("냉난방면적 (평 환산)", "=B{0}/3.30579".format(R_AREA), "평",
         area.get("설치계획서_냉난방면적_평"), '=IF(ABS(B{r}-D{r})<=1,"✅","⚠️")',
         "㎡ ÷ 3.30579"),
        ("연면적 대비 냉난방면적 비율", "=B{0}/D{1}".format(R_AREA, R_AREA + 5), "%",
         area.get("연면적_m2"), NA, "D열=연면적(㎡), 참고지표"),
    ]
    for name, val, u, comp, judge, note in area_rows:
        if comp is None:                       # 대조값이 없으면 검증 불가
            val, comp, judge = NA, NA, NA
        _row(ws, r, [name, val, u, comp, judge.format(r=r) if judge != NA else NA], ncols=5)
        ws.cell(r, 6, note).font = NOTE_F
        r += 1
    ws.cell(R_AREA + 5, 2).number_format = "0.0%"

    # [3] 최대부하 및 최대부하율
    ws["A18"] = "[3] 최대부하 및 최대부하율"
    ws["A18"].font = H1_F
    _header(ws, R_LOAD_HDR, ["항목", "계산값(수식)", "단위", "설치계획서 기재값", "판정"])
    rt = "=B{0}/{1}"
    load_rows = [
        ("냉방 최대부하 합계", "=SUM(D{0}:D{1})".format(R_SPACE, SPACE_LAST), unit,
         load.get("계산서_냉방부하합계"), '=IF(ABS(B{r}-D{r})<=1,"✅","⚠️")',
         "D열=계산서 기재 합계 (전사 누락 검증)"),
        ("냉방 최대부하율", "=B{0}/B{1}".format(R_LOAD, R_AREA), "%s/㎡" % unit, NA, NA,
         "냉방 최대부하 ÷ 냉방면적"),   # '='로 시작하는 표시용 문자열 금지 (엑셀이 수식으로 저장 후 제거)
        ("냉방 최대부하(RT)", rt.format(R_LOAD, div), "RT",
         load.get("설치계획서_냉방최대부하_RT"), '=IF(ABS(B{r}-D{r})<=D{r}*0.05,"✅","❌")',
         "1 RT = %s %s" % ("{:,}".format(div) if div else "-", unit)),
        ("난방 최대부하 합계", "=SUM(G{0}:G{1})".format(R_SPACE, SPACE_LAST), unit,
         load.get("계산서_난방부하합계"), '=IF(ABS(B{r}-D{r})<=1,"✅","⚠️")',
         "D열=계산서 기재 합계"),
        ("난방 최대부하율", "=B{0}/B{1}".format(R_LOAD + 3, R_AREA + 1), "%s/㎡" % unit, NA, NA,
         "난방 최대부하 ÷ 난방면적"),   # '='로 시작하는 표시용 문자열 금지
        ("난방 최대부하(RT)", rt.format(R_LOAD + 3, div), "RT",
         load.get("설치계획서_난방최대부하_RT"),
         '=IF(ABS(B{r}-D{r})<=MAX(1,D{r}*0.05),"✅","❌")', "설치계획서 기재값과 대조"),
    ]
    r = R_LOAD
    heating = load.get("계산서_난방부하합계") is not None
    for i, (name, val, u, plan, judge, note) in enumerate(load_rows):
        if i >= 3 and not heating:             # 냉열전용: 난방 3행은 해당없음
            val, plan, judge = NA, NA, NA
        if plan is None:
            plan, judge = NA, NA
        _row(ws, r, [name, val, u, plan, judge.format(r=r) if judge != NA else NA], ncols=5)
        ws.cell(r, 6, note).font = NOTE_F
        r += 1

    # [4] 공간별 면적 및 냉난방부하
    ws.cell(R_SPACE_TITLE, 1, "[4] 공간별 면적 및 냉난방부하").font = H1_F
    _header(ws, R_SPACE_HDR,
            ["공간(실/존/계통)", "실수", "냉방면적(㎡)", "냉방부하(%s)" % unit, "냉방부하율(W/㎡)",
             "난방면적(㎡)", "난방부하(%s)" % unit, "난방부하율(W/㎡)",
             "단위부하(기재)", "부하 재계산(수식)", "재계산 판정"])
    r = R_SPACE
    for s in spec["spaces"]:
        ca, ha = s.get("냉방면적"), s.get("난방면적")
        ws.cell(r, 1, s["공간"])
        ws.cell(r, 2, s.get("실수", NA))
        ws.cell(r, 3, ca if ca is not None else "— (미기재)")
        ws.cell(r, 4, s.get("냉방부하", NA))
        ws.cell(r, 5, "=D{0}/C{0}".format(r) if ca is not None else NA)
        ws.cell(r, 6, ha if ha is not None else "— (미기재)")
        ws.cell(r, 7, s.get("난방부하", NA))
        ws.cell(r, 8, "=G{0}/F{0}".format(r) if ha is not None else NA)
        # 면적 × 단위부하 방식이면 행 단위 재계산 검증
        up = s.get("단위부하")
        if up is not None and ca is not None:
            ws.cell(r, 9, up)
            ws.cell(r, 10, "=C{0}*I{0}".format(r))
            ws.cell(r, 11, '=IF(ABS(J{0}-D{0})<=MAX(1,D{0}*0.01),"✅","⚠️")'.format(r))
        else:
            for c in (9, 10, 11):
                ws.cell(r, c, NA)
        for c in range(1, 12):
            ws.cell(r, c).border = BOX
        for c in (3, 4, 6, 7, 10):
            ws.cell(r, c).number_format = "#,##0"
        for c in (5, 8):
            ws.cell(r, c).number_format = "0.0"
        r += 1

    for rr in list(range(R_AREA, R_AREA + 5)) + [R_LOAD, R_LOAD + 2, R_LOAD + 3, R_LOAD + 5]:
        ws.cell(rr, 2).number_format = "#,##0.0"
        ws.cell(rr, 4).number_format = "#,##0.0"
    for rr in (R_LOAD + 1, R_LOAD + 4):
        ws.cell(rr, 2).number_format = "0.0"

    if notes:
        write_notes(ws, spec, start_row=r + 1)
    for col, w in zip("ABCDEFGHIJK", (30, 26, 12, 22, 46, 16, 16, 16, 14, 16, 12)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"


def write_notes(ws, spec, start_row):
    """[5] 확인사항"""
    notes = spec.get("notes") or []
    r = start_row
    ws.cell(r, 1, "[5] 확인사항").font = H1_F
    r += 1
    _header(ws, r, ["번호", "항목", "부하계산서", "설치계획서", "확인요청 내용"])
    r += 1
    for i, n in enumerate(notes, 1):
        _row(ws, r, [i, n["항목"], n.get("부하계산서", NA), n.get("설치계획서", NA), n["확인요청"]],
             ncols=5)
        for c in range(1, 6):
            ws.cell(r, c).alignment = Alignment(vertical="top", wrap_text=True)
        r += 1
    for col, w in zip("ABCDE", (30, 26, 12, 22, 46)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"


def build_rooms(wb, detail):
    """실별 상세표 전량 전사 + 열별 합계 대조 검증. (판정결과 리스트를 돌려준다)"""
    ws = replace_sheet(wb, detail.get("sheet_name", "실별부하집계"))
    ws["A1"] = detail["title"]
    ws["A1"].font = TITLE_F
    if detail.get("subtitle"):
        ws["A2"] = detail["subtitle"]
        ws["A2"].font = NOTE_F

    cols = detail["columns"]                     # 냉방 시각 열 이름들
    heat = detail.get("heating_column", "난방(W)")
    _header(ws, 3, ["No.", "실명", "Q'ty"] + cols + [heat, "합계반영", "비고"])
    ncol = 3 + len(cols) + 1                     # 난방 열까지의 열 번호
    col_flag, col_note = ncol + 1, ncol + 2

    r = 4
    for row in detail["rows"]:
        no, name, qty, values, heating, flag, note = (
            row["no"], row["실명"], row.get("qty", 1), row["냉방"], row.get("난방"),
            row.get("합계반영", 1), row.get("비고", ""))
        ws.cell(r, 1, no)
        ws.cell(r, 2, name)
        ws.cell(r, 3, qty)
        for i, v in enumerate(values):
            ws.cell(r, 4 + i, v if v is not None else "-")
        ws.cell(r, ncol, heating if heating is not None else "-")
        ws.cell(r, col_flag, flag)
        ws.cell(r, col_note, note)
        for c in range(1, col_note + 1):
            ws.cell(r, c).border = BOX
            if 4 <= c <= ncol:
                ws.cell(r, c).number_format = "#,##0"
                ws.cell(r, c).alignment = Alignment(horizontal="right")
            if not flag:
                ws.cell(r, c).fill = DUP_B
        r += 1
    last = r - 1

    # 검증 블록
    r += 1
    ws.cell(r, 1, "[검증] 실별 합계 vs 계산서 기재 Grand Total").font = H1_F
    r += 1
    _header(ws, r, ["구분"] + cols + [heat.split("(")[0]])
    r += 1
    scale = 1000.0 if detail["grand_total"].get("unit", "kW") == "kW" else 1.0
    flag_col = chr(ord("A") + col_flag - 1)
    ws.cell(r, 1, "실별 합계 (합계반영 행만, %s)" % detail["grand_total"].get("unit", "kW"))
    for i in range(len(cols) + 1):
        col = chr(ord("D") + i)
        # SUMIF는 합계범위의 텍스트('-')를 무시한다 (SUMPRODUCT는 #VALUE! 발생)
        ws.cell(r, 2 + i, "=SUMIF(${f}$4:${f}${l},1,{c}4:{c}{l}){s}".format(
            f=flag_col, l=last, c=col, s="/1000" if scale == 1000.0 else ""))
    row_sum = r

    r += 1
    ws.cell(r, 1, "계산서 기재 Grand Total (%s)" % detail["grand_total"].get("unit", "kW"))
    for i, v in enumerate(detail["grand_total"]["values"] + [detail["grand_total"]["heating"]]):
        ws.cell(r, 2 + i, v)
    row_grand = r

    r += 1
    tol = detail["grand_total"].get("tolerance", 0.2)
    ws.cell(r, 1, "판정 (±%s)" % tol)
    for i in range(len(cols) + 1):
        col = chr(ord("B") + i)
        ws.cell(r, 2 + i, '=IF(ABS({c}{s}-{c}{g})<={t},"✅","⚠️")'.format(
            c=col, s=row_sum, g=row_grand, t=tol))

    r += 1
    ws.cell(r, 1, "실수 (Q'ty) 합계 / 기재")
    ws.cell(r, 2, "=SUMIF(${f}$4:${f}${l},1,$C$4:$C${l})".format(f=flag_col, l=last))
    ws.cell(r, 3, detail["grand_total"]["qty"])
    ws.cell(r, 4, '=IF(B{0}=C{0},"✅","⚠️")'.format(r))
    row_qty = r

    for rr in range(row_sum - 1, row_qty + 1):
        for c in range(1, len(cols) + 3):
            ws.cell(rr, c).border = BOX
    for rr in (row_sum, row_grand):
        ws.cell(rr, 1).font = Font(bold=True)
        for c in range(1, len(cols) + 3):
            ws.cell(rr, c).fill = SUM_B
            if c >= 2:
                ws.cell(rr, c).number_format = "#,##0.0"

    if detail.get("footnote"):
        r += 2
        ws.cell(r, 1, "※ " + detail["footnote"]).font = NOTE_F

    widths = [8, 26, 6] + [11] * len(cols) + [12, 10, 44]
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(ord("A") + i)].width = w
    ws.freeze_panes = "D4"
    ws.auto_filter.ref = "A3:%s%d" % (chr(ord("A") + col_note - 1), last)

    return verify_rooms(detail, scale, tol)


def verify_rooms(detail, scale, tol):
    """빌드 단계 자체 검증 — 열별 합계가 기재 Grand Total과 맞는지 파이썬에서도 확인."""
    out = []
    gt = detail["grand_total"]
    targets = gt["values"] + [gt["heating"]]
    n = len(detail["columns"])
    for i in range(n + 1):
        s = 0.0
        for row in detail["rows"]:
            if not row.get("합계반영", 1):
                continue
            v = row["냉방"][i] if i < n else row.get("난방")
            s += v or 0
        s /= scale
        label = (detail["columns"] + [detail.get("heating_column", "난방")])[i]
        out.append((label, s, targets[i], abs(s - targets[i]) <= tol))
    qty = sum(r.get("qty", 1) for r in detail["rows"] if r.get("합계반영", 1))
    out.append(("Q'ty", qty, gt["qty"], qty == gt["qty"]))
    return out


def main():
    ap = argparse.ArgumentParser(description="건물부하계산서 검토 엑셀 빌더 (Phase 2 / v13)")
    ap.add_argument("spec", help="판독값 JSON 경로")
    ap.add_argument("-o", "--output", help="출력 xlsx (기본: JSON의 output)")
    ap.add_argument("--into", help="기존 분석 워크북에 시트를 끼워 넣는다 (단일 파일 산출 — v17 기본)")
    args = ap.parse_args()

    with open(args.spec, encoding="utf-8") as f:
        spec = json.load(f)
    title = spec.get("sheet_name", "건물부하계산서")

    if args.into:                        # 단일 파일 모드 — 12시트 워크북에 병합
        out = args.into
        wb = open_or_new(out)
        ws = replace_sheet(wb, title)
    else:                                # 단독 검증 파일 모드
        out = args.output or spec.get("output") or "건물부하계산서.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = title
    build_main(ws, spec, notes=not args.into)

    checks = []
    if spec.get("room_detail") and spec["presence"].get("유무") == "유":
        checks = build_rooms(wb, spec["room_detail"])

    if args.into:
        alias = {}
        if spec.get("room_detail"):
            alias[spec["room_detail"].get("sheet_name", "실별부하집계")] = "실별부하집계"
        order_sheets(wb, alias)

    save_deterministic(wb, out)          # 재실행 시 바이트 동일 보장(R8)
    print("saved:", out)
    print("sha256:", sha256(out))

    if checks:
        print("\n[실별 합계 대조]")
        bad = 0
        for label, s, g, ok in checks:
            print("  %-8s 전사 %10.1f | 기재 %10.1f | %s" % (label, s, g, "✅" if ok else "⚠️"))
            bad += (not ok)
        if bad:
            print("\n⚠️ %d개 열이 불일치한다. 원문을 재판독하거나 중복·소계 행의 "
                  "'합계반영'을 0으로 처리하라." % bad)
            return 1
        print("  → 전 열 일치. 전사 정확도 확인됨.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
